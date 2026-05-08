"""Workbook I/O: loading transcription workbooks and writing scored outputs.

Keeps all openpyxl and file-system concerns out of the core scoring logic.
"""

from __future__ import annotations

import re
from pathlib import Path
from shutil import copyfile
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from autoeit.utils.paths import ensure_parent_dir

# --------------------------------------------------------------------------
# Styling constants for the output workbook
# --------------------------------------------------------------------------

SCORE_FILLS = {
    4: PatternFill(fill_type="solid", fgColor="C6EFCE"),
    3: PatternFill(fill_type="solid", fgColor="FFEB9C"),
    2: PatternFill(fill_type="solid", fgColor="FFCC99"),
    1: PatternFill(fill_type="solid", fgColor="FFC7CE"),
    0: PatternFill(fill_type="solid", fgColor="FF0000"),
}
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# --------------------------------------------------------------------------
# Sheet-name parsing
# --------------------------------------------------------------------------

_SHEET_PATTERNS = [
    re.compile(r"^(?P<pid>\d+)_(?P<version>v[AB])$"),
    re.compile(r"^(?P<pid>\d+)-(?P<version>\d[AB])$"),
]


def _parse_sheet_name(name: str) -> tuple[str, str]:
    """Extract (participant_id, version) from a worksheet tab name.

    Returns (name, "") if the name doesn't match any known pattern — handles
    summary tabs and one-off sheet naming conventions gracefully.
    """
    for pat in _SHEET_PATTERNS:
        m = pat.match(name)
        if m:
            return m.group("pid"), m.group("version")
    return name, ""


def _is_scoring_sheet(ws: Worksheet) -> bool:
    return (
        ws.cell(row=1, column=1).value == "Sentence"
        and ws.cell(row=1, column=2).value == "Stimulus"
    )


def _last_header_column(ws: Worksheet) -> int:
    for col in range(ws.max_column, 0, -1):
        v = ws.cell(row=1, column=col).value
        if v is not None and str(v).strip():
            return col
    return 1


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------


def validate_schema(filepath: str | Path) -> list[str]:
    """Return a list of schema errors for the workbook at *filepath*.

    An empty list means the workbook looks usable.
    """
    errors: list[str] = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        errors.append(f"Cannot open workbook: {exc}")
        return errors

    try:
        scoring = [n for n in wb.sheetnames if _is_scoring_sheet(wb[n])]
        if not scoring:
            errors.append(
                "No scoring sheets found — expected at least one sheet with "
                "'Sentence' in A1 and 'Stimulus' in B1."
            )
            return errors

        for name in scoring:
            ws = wb[name]
            has_rows = False
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if isinstance(row[0], int):
                    has_rows = True
                    if len(row) < 3:
                        errors.append(
                            f"Sheet '{name}' row {i}: fewer than 3 columns "
                            "(need Sentence, Stimulus, Transcription)."
                        )
                    break
            if not has_rows:
                errors.append(f"Sheet '{name}': no sentence rows found (integer expected in column A).")
    finally:
        wb.close()

    return errors


# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------


def load_workbook(filepath: str | Path) -> pd.DataFrame:
    """Parse all scoring sheets into a single flat DataFrame.

    Raises ValueError if no scoring sheets are found or if they're empty.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    records: list[dict[str, object]] = []
    found_any = False

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not _is_scoring_sheet(ws):
            continue
        found_any = True
        pid, version = _parse_sheet_name(sheet_name)

        for row in ws.iter_rows(min_row=2, values_only=True):
            sentence_id = row[0]
            if not isinstance(sentence_id, int):
                continue
            records.append({
                "sheet_name": sheet_name,
                "participant_id": pid,
                "version": version,
                "sentence_id": sentence_id,
                "stimulus": row[1],
                "transcription": row[2] if len(row) >= 3 else None,
                "human_score": row[3] if len(row) >= 4 else None,
            })

    if not found_any:
        raise ValueError("Workbook has no scoring sheets (Sentence/Stimulus headers not found).")
    if not records:
        raise ValueError("Scoring sheets found but contain no sentence rows.")

    return (
        pd.DataFrame(records)
        .sort_values(["sheet_name", "sentence_id"])
        .reset_index(drop=True)
    )


# --------------------------------------------------------------------------
# Saving
# --------------------------------------------------------------------------


def save_workbook(
    frame: pd.DataFrame,
    *,
    source_path: str | Path,
    out_path: str | Path,
) -> Path:
    """Write an annotated copy of *source_path* to *out_path*.

    Copies the original workbook, then appends AutoEIT_Score and Rationale
    columns to every scoring sheet plus a consolidated AutoEIT_Summary tab.
    """
    out = ensure_parent_dir(out_path)
    copyfile(source_path, out)

    wb = openpyxl.load_workbook(out)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not _is_scoring_sheet(ws):
            continue

        lookup = {
            int(r["sentence_id"]): (int(r["auto_score"]), r["rationale"])
            for _, r in frame[frame["sheet_name"] == sheet_name].iterrows()
        }
        base = _last_header_column(ws)
        sc, rc = base + 1, base + 2

        for col, label in ((sc, "AutoEIT_Score"), (rc, "Rationale")):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for row_idx in range(2, ws.max_row + 1):
            sid = ws.cell(row=row_idx, column=1).value
            if not isinstance(sid, int) or sid not in lookup:
                continue
            score, rationale = lookup[sid]
            ws.cell(row=row_idx, column=sc, value=score).fill = SCORE_FILLS[score]
            ws.cell(row=row_idx, column=sc).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=rc, value=rationale).alignment = Alignment(wrap_text=True)

    # Build a consolidated summary sheet
    if "AutoEIT_Summary" in wb.sheetnames:
        del wb["AutoEIT_Summary"]
    summary = wb.create_sheet("AutoEIT_Summary", 0)

    headers = [
        "Sheet", "Participant", "Version", "Sentence", "Stimulus",
        "Transcription", "Human Score", "AutoEIT Score", "Score Diff", "Rationale",
    ]
    col_widths = [18, 14, 10, 10, 50, 55, 13, 13, 11, 60]
    for i, h in enumerate(headers, 1):
        cell = summary.cell(row=1, column=i, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        summary.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]

    for row_idx, (_, row) in enumerate(frame.iterrows(), 2):
        diff: int | str = ""
        if pd.notna(row.get("human_score")):
            diff = int(row["auto_score"] - row["human_score"])
        vals: list[Any] = [
            row["sheet_name"], row["participant_id"], row["version"],
            int(row["sentence_id"]), row["stimulus"], row["transcription"],
            int(row["human_score"]) if pd.notna(row["human_score"]) else "",
            int(row["auto_score"]), diff, row["rationale"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = summary.cell(row=row_idx, column=ci, value=v)
            cell.alignment = Alignment(wrap_text=True)
            if ci == 8:  # score column
                cell.fill = SCORE_FILLS[int(row["auto_score"])]
                cell.alignment = Alignment(horizontal="center")

    summary.freeze_panes = "A2"
    wb.save(out)
    return out


def save_csv_outputs(
    frame: pd.DataFrame,
    *,
    scores_path: str | Path,
    downgrades_path: str | Path,
) -> tuple[Path, Path]:
    """Write the main scores CSV and an audit log of ambiguous downgrades."""
    sp = ensure_parent_dir(scores_path)
    dp = ensure_parent_dir(downgrades_path)

    frame[[
        "sheet_name", "participant_id", "version", "sentence_id",
        "stimulus", "transcription", "human_score", "auto_score",
        "ambiguous_downgraded", "rationale",
    ]].to_csv(sp, index=False)

    frame[frame["ambiguous_downgraded"]][[
        "sheet_name", "participant_id", "version", "sentence_id",
        "stimulus", "transcription", "human_score", "auto_score", "rationale",
    ]].to_csv(dp, index=False)

    return sp, dp
