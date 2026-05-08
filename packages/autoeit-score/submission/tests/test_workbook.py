"""Tests for autoeit.services.workbook — loading and schema validation.

Uses in-memory workbooks rather than real fixtures to keep tests fast and
self-contained.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from autoeit.services.workbook import load_workbook, validate_schema, _parse_sheet_name


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(sheets: dict[str, list[tuple]]) -> Path:
    f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    path = Path(f.name)
    f.close()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    return path


_VALID_ROWS = [
    ("Sentence", "Stimulus", "Transcription", "Human Score"),
    (1, "El gato corre", "el gato corre", 4),
    (2, "La niña lee", "la nina lee", 3),
]


# ---------------------------------------------------------------------------
# _parse_sheet_name
# ---------------------------------------------------------------------------

class TestParseSheetName:
    def test_underscore_vA(self):
        pid, ver = _parse_sheet_name("42_vA")
        assert pid == "42" and ver == "vA"

    def test_dash_format(self):
        pid, ver = _parse_sheet_name("12-1B")
        assert pid == "12" and ver == "1B"

    def test_unrecognised(self):
        pid, ver = _parse_sheet_name("Summary")
        assert pid == "Summary" and ver == ""


# ---------------------------------------------------------------------------
# validate_schema
# ---------------------------------------------------------------------------

class TestValidateSchema:
    def test_valid_workbook(self):
        path = _make_workbook({"42_vA": _VALID_ROWS})
        errors = validate_schema(path)
        assert errors == []

    def test_no_scoring_sheets(self):
        path = _make_workbook({"Cover": [("Title", "AutoEIT")]})
        errors = validate_schema(path)
        assert len(errors) > 0

    def test_nonexistent_file(self):
        errors = validate_schema("/tmp/does_not_exist_xyz.xlsx")
        assert len(errors) > 0


# ---------------------------------------------------------------------------
# load_workbook
# ---------------------------------------------------------------------------

class TestLoadWorkbook:
    def test_returns_expected_columns(self):
        path = _make_workbook({"42_vA": _VALID_ROWS})
        df = load_workbook(path)
        for col in ("sheet_name", "participant_id", "version", "sentence_id",
                    "stimulus", "transcription", "human_score"):
            assert col in df.columns

    def test_row_count(self):
        path = _make_workbook({"42_vA": _VALID_ROWS})
        df = load_workbook(path)
        assert len(df) == 2  # two data rows

    def test_no_scoring_sheets_raises(self):
        path = _make_workbook({"Cover": [("Info", "Data")]})
        with pytest.raises(ValueError, match="no scoring sheets"):
            load_workbook(path)

    def test_participant_id_extracted(self):
        path = _make_workbook({"7_vB": _VALID_ROWS})
        df = load_workbook(path)
        assert (df["participant_id"] == "7").all()

    def test_missing_human_score_column(self):
        rows = [
            ("Sentence", "Stimulus", "Transcription"),
            (1, "El gato corre", "el gato corre"),
        ]
        path = _make_workbook({"1_vA": rows})
        df = load_workbook(path)
        assert df["human_score"].isna().all()
