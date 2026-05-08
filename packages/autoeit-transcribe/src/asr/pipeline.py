"""Task I transcription pipeline for AutoEIT."""

from __future__ import annotations

from dataclasses import dataclass
import os
from pathlib import Path
from shutil import copyfile
import re
from typing import Callable, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.io.workbooks import ensure_parent_dir, last_populated_header_column
from src.postprocess.normalization import collapse_whitespace, normalize_transcription_text
from src.postprocess.hallucination import cleanup_transcription as _cleanup_transcription
from src.asr.model import build_whisper_transcriber as _build_whisper_transcriber

EXPECTED_SENTENCE_COUNT = 30

HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class AlignmentError(ValueError):
    """Raised when ASR segments cannot be deterministically aligned."""


@dataclass(frozen=True)
class SegmentChunk:
    start: float
    end: float
    text: str


@dataclass(frozen=True)
class ParticipantJob:
    sheet_name: str
    participant_id: str
    audio_filename: str
    prompts: list[tuple[int, str]]


@dataclass(frozen=True)
class TranscriptRow:
    sentence_id: int
    stimulus: str
    raw_transcription: str
    normalized_transcription: str
    notes: str


@dataclass(frozen=True)
class ParticipantTranscript:
    sheet_name: str
    participant_id: str
    audio_filename: str
    rows: list[TranscriptRow]


Transcriber = Callable[[Path], list[SegmentChunk]]


def cleanup_transcription(text: str) -> str:
    """Normalize and remove hallucinated transcription content.

    This delegates to the shared postprocessing implementation to avoid
    divergence of hallucination patterns and behavior.
    """
    return _cleanup_transcription(text)


def build_whisper_transcriber(
    *,
    model_size: str,
    device: str,
    compute_type: str,
    language: str,
) -> Transcriber:
    """Build a Whisper-based transcriber using the shared ASR model helper.

    Delegates to ``src.asr.model.build_whisper_transcriber`` to keep model
    configuration and behavior consistent across the codebase.
    """
    return _build_whisper_transcriber(
        model_size=model_size,
        device=device,
        compute_type=compute_type,
        language=language,
    )


def parse_prompt_jobs(
    prompt_workbook: str | Path,
    *,
    expected_count: int = EXPECTED_SENTENCE_COUNT,
    only_sheet: str | None = None,
) -> list[ParticipantJob]:
    workbook = openpyxl.load_workbook(prompt_workbook, data_only=True)
    jobs: list[ParticipantJob] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name == "Info":
            continue
        if only_sheet and sheet_name != only_sheet:
            continue
        worksheet = workbook[sheet_name]
        if worksheet.cell(row=1, column=1).value != "Sentence":
            continue
        prompts: list[tuple[int, str]] = []
        audio_filename = ""
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            sentence_id = row[0]
            stimulus = row[1]
            candidate_audio = row[6] if len(row) >= 7 else None
            if isinstance(sentence_id, int) and isinstance(stimulus, str):
                prompts.append((sentence_id, stimulus))
            if not audio_filename and isinstance(candidate_audio, str):
                audio_filename = candidate_audio
        if not prompts:
            continue
        if len(prompts) != expected_count:
            raise AlignmentError(
                f"Sheet {sheet_name} has {len(prompts)} prompts; expected {expected_count}."
            )
        if not audio_filename:
            match = re.match(r"^(?P<pid>\d+)-(?P<version>\d[A-Z])$", sheet_name)
            if match:
                participant = match.group("pid").zfill(6)
                version = match.group("version")
                audio_filename = f"{participant}_EIT-{version}.mp3"
        if not audio_filename:
            raise FileNotFoundError(f"Sheet {sheet_name} does not specify an audio filename.")
        jobs.append(
            ParticipantJob(
                sheet_name=sheet_name,
                participant_id=sheet_name,
                audio_filename=audio_filename,
                prompts=prompts,
            )
        )
    if not jobs:
        raise ValueError("No participant sheets were found in the Task I workbook.")
    return jobs


def _merge_candidate_key(
    left: SegmentChunk,
    right: SegmentChunk,
    index: int,
) -> tuple[float, int, int]:
    gap = max(0.0, right.start - left.end)
    token_count = len(left.text.split()) + len(right.text.split())
    return (gap, token_count, index)


def align_segments_to_prompts(
    segments: Iterable[SegmentChunk],
    *,
    expected_count: int = EXPECTED_SENTENCE_COUNT,
) -> list[str]:
    groups = [segment for segment in segments if cleanup_transcription(segment.text)]
    if len(groups) < expected_count:
        deficit = expected_count - len(groups)
        while len(groups) < expected_count:
            candidates = [
                (index, segment)
                for index, segment in enumerate(groups)
                if len(segment.text.split()) >= 6
            ]
            if not candidates:
                break
            split_index, segment = max(
                candidates,
                key=lambda item: ((item[1].end - item[1].start), len(item[1].text.split())),
            )
            tokens = segment.text.split()
            midpoint = len(tokens) // 2
            left_text = cleanup_transcription(" ".join(tokens[:midpoint]))
            right_text = cleanup_transcription(" ".join(tokens[midpoint:]))
            if not left_text or not right_text:
                break
            midpoint_time = segment.start + (segment.end - segment.start) / 2.0
            groups[split_index : split_index + 1] = [
                SegmentChunk(start=segment.start, end=midpoint_time, text=left_text),
                SegmentChunk(start=midpoint_time, end=segment.end, text=right_text),
            ]
        if len(groups) < expected_count:
            raise AlignmentError(
                f"ASR produced {len(groups)} non-empty segments after fallback splitting; expected {expected_count}."
            )
    while len(groups) > expected_count:
        merge_index = min(
            range(len(groups) - 1),
            key=lambda idx: _merge_candidate_key(groups[idx], groups[idx + 1], idx),
        )
        left = groups[merge_index]
        right = groups[merge_index + 1]
        groups[merge_index : merge_index + 2] = [
            SegmentChunk(
                start=left.start,
                end=right.end,
                text=cleanup_transcription(f"{left.text} {right.text}"),
            )
        ]
    return [cleanup_transcription(group.text) for group in groups]


def transcribe_participant(
    job: ParticipantJob,
    *,
    audio_dir: str | Path,
    transcriber: Transcriber,
    expected_count: int = EXPECTED_SENTENCE_COUNT,
) -> ParticipantTranscript:
    audio_path = Path(audio_dir) / job.audio_filename
    if not audio_path.exists():
        raise FileNotFoundError(f"Missing audio file: {audio_path}")
    aligned = align_segments_to_prompts(
        transcriber(audio_path),
        expected_count=expected_count,
    )
    rows: list[TranscriptRow] = []
    for (sentence_id, stimulus), raw_text in zip(job.prompts, aligned, strict=True):
        rows.append(
            TranscriptRow(
                sentence_id=sentence_id,
                stimulus=stimulus,
                raw_transcription=raw_text,
                normalized_transcription=normalize_transcription_text(raw_text),
                notes="",
            )
        )
    return ParticipantTranscript(
        sheet_name=job.sheet_name,
        participant_id=job.participant_id,
        audio_filename=job.audio_filename,
        rows=rows,
    )


def write_output_workbook(
    transcripts: list[ParticipantTranscript],
    *,
    source_workbook: str | Path,
    output_workbook: str | Path,
) -> Path:
    output_path = ensure_parent_dir(output_workbook)
    copyfile(source_workbook, output_path)
    workbook = openpyxl.load_workbook(output_path)

    if "AutoEIT_Task1_Summary" in workbook.sheetnames:
        del workbook["AutoEIT_Task1_Summary"]
    summary = workbook.create_sheet("AutoEIT_Task1_Summary", 0)
    summary_headers = [
        "Participant",
        "Audio File",
        "Sentence",
        "Stimulus",
        "Raw Transcription",
        "Normalized Transcription",
        "Notes",
    ]
    for col_index, header in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col_index, width in enumerate([16, 24, 10, 48, 48, 40, 20], start=1):
        summary.column_dimensions[get_column_letter(col_index)].width = width

    summary_row = 2
    for transcript in transcripts:
        worksheet = workbook[transcript.sheet_name]
        start_col = last_populated_header_column(worksheet)
        worksheet.cell(row=1, column=max(start_col, 3), value="Transcription")
        worksheet.cell(row=1, column=max(start_col + 1, 4), value="Normalized transcription")
        worksheet.cell(row=1, column=max(start_col + 2, 5), value="Notes")
        for col in (max(start_col, 3), max(start_col + 1, 4), max(start_col + 2, 5)):
            header_cell = worksheet.cell(row=1, column=col)
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.alignment = Alignment(horizontal="center")

        for row_index, row in enumerate(transcript.rows, start=2):
            worksheet.cell(row=row_index, column=max(start_col, 3), value=row.raw_transcription)
            worksheet.cell(
                row=row_index,
                column=max(start_col + 1, 4),
                value=row.normalized_transcription,
            )
            worksheet.cell(row=row_index, column=max(start_col + 2, 5), value=row.notes)

            values = [
                transcript.participant_id,
                transcript.audio_filename,
                row.sentence_id,
                row.stimulus,
                row.raw_transcription,
                row.normalized_transcription,
                row.notes,
            ]
            for col_index, value in enumerate(values, start=1):
                summary.cell(row=summary_row, column=col_index, value=value)
                summary.cell(row=summary_row, column=col_index).alignment = Alignment(
                    wrap_text=True
                )
            summary_row += 1

    summary.freeze_panes = "A2"
    workbook.save(output_path)
    return output_path


def run_batch(
    *,
    audio_dir: str | Path,
    prompt_workbook: str | Path,
    output_workbook: str | Path,
    expected_count: int = EXPECTED_SENTENCE_COUNT,
    model_size: str = "large-v3",
    device: str = "cpu",
    compute_type: str = "int8",
    language: str = "es",
    only_sheet: str | None = None,
    transcriber: Transcriber | None = None,
) -> list[ParticipantTranscript]:
    jobs = parse_prompt_jobs(prompt_workbook, expected_count=expected_count, only_sheet=only_sheet)
    job_transcriber = transcriber or build_whisper_transcriber(
        model_size=model_size,
        device=device,
        compute_type=compute_type,
        language=language,
    )
    transcripts = [
        transcribe_participant(
            job,
            audio_dir=audio_dir,
            transcriber=job_transcriber,
            expected_count=expected_count,
        )
        for job in jobs
    ]
    write_output_workbook(
        transcripts,
        source_workbook=prompt_workbook,
        output_workbook=output_workbook,
    )
    return transcripts
